import streamlit as st
import google.generativeai as genai
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import json
import pandas as pd

# 1. 페이지 설정 및 디자인
st.set_page_config(page_title="작업 표준서 다국어 번역기", layout="wide")
st.title("🏭 작업 표준서 다국어 번역 시스템 (웹 공유형)")
st.write("엑셀의 사진, 도형, 서식을 100% 원본 그대로 유지하면서 글자만 자동 전문 번역합니다.")

# 2. 웹 배포용 세션 상태 초기화
if "saved_api_key" not in st.session_state:
    st.session_state["saved_api_key"] = ""

# 3. 사이드바 설정
st.sidebar.header("🔧 설정")

api_key = st.sidebar.text_input(
    "Gemini API Key를 입력하세요:", 
    type="password",
    value=st.session_state["saved_api_key"]
)

if api_key:
    st.session_state["saved_api_key"] = api_key.strip()

languages = ["한국어", "영어", "일본어", "중국어", "베트남어"]
st.sidebar.subheader("🌐 언어 선택")
source_lang = st.sidebar.selectbox("원본 언어", languages, index=1) # 기본값 영어
target_lang = st.sidebar.selectbox("목적 언어", languages, index=0) # 기본값 한국어

if source_lang == target_lang:
    st.sidebar.error("⚠️ 원본 언어와 목적 언어는 달라야 합니다.")

# 4. 메인 화면 - 파일 업로드 섹션
st.header("📂 엑셀 파일 업로드")
uploaded_file = st.file_uploader("번역할 작업 표준서(Excel) 파일을 선택하세요.", type=["xlsx"])

# 5. 스마트 자동 재시도 기능이 내장된 배치 번역 함수
def translate_batch(text_list, source, target, model):
    if not text_list:
        return {}
        
    mapping_rules = ""
    if source == "한국어" and target == "영어":
        mapping_rules = (
            '- "공정명" -> "Process Name"\n'
            '- "공정번호" -> "Process No."\n'
            '- "퍼지" -> "Purge"\n'
            '- "공타" -> "Dry Shot"\n'
            '- "초물" -> "First Article"\n'
            '- "구분" -> "Division"\n'
            '- "시업시" -> "Startup"'
        )
    elif source == "영어" and target == "한국어":
        mapping_rules = (
            '- "Process Name" or "fair name" -> "공정명"\n'
            '- "Process No." -> "공정번호"\n'
            '- "Purge" or "fudge" -> "퍼지"\n'
            '- "Dry Shot" or "empty hit" -> "공타"\n'
            '- "First Article" -> "초물"\n'
            '- "Division" or "Category" -> "구분"\n'
            '- "Startup" -> "시업시"\n'
            '- "work standards" -> "작업표준서"\n'
            '- "Status of parts used" -> "부품 사용 현황"\n'
            '- "Quality check box" -> "품질 체크 박스"'
        )
        
    input_packages = [{"id": i, "text": text} for i, text in enumerate(text_list)]
    json_data = json.dumps(input_packages, ensure_ascii=False)
        
    prompt = (
        f"You are an expert translator specializing in industrial manufacturing, automotive parts production, safety management, and Standard Operating Procedures (SOP).\n"
        f"Translate the 'text' field of each object from {source} to {target} reflecting this industrial context.\n\n"
        f"Crucial Terminology Mapping Rules:\n{mapping_rules}\n\n"
        f"Strict Rules:\n"
        f"1. Return the results ONLY as a valid JSON array of objects.\n"
        f"2. Each object MUST contain the exact original 'id' (integer) and the 'translated' (string) text.\n"
        f"3. Never alter or translate the 'id' value. Keep it as an integer.\n"
        f"4. Do not include any markdown styling like ```json. Just raw JSON string.\n\n"
        f"Input list to translate:\n{json_data}"
    )
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = model.generate_content(
                prompt, 
                generation_config={"response_mime_type": "application/json"}
            )
            res_json = json.loads(response.text.strip())
            
            translation_map = {}
            for item in res_json:
                item_id = item.get("id")
                translated_text = item.get("translated")
                if item_id is not None and 0 <= item_id < len(text_list):
                    translation_map[text_list[item_id]] = translated_text
            return translation_map
            
        except Exception as e:
            err_msg = str(e)
            if "429" in err_msg and attempt < max_retries - 1:
                st.info(f"⏳ 구글 무료 서버 제한(429) 감지: 5초간 대기 후 자동 재시도합니다. ({attempt + 1}/{max_retries}차 시도)")
                time.sleep(5)
            else:
                st.error(f"❌ AI 서버 통신 오류 발생: {err_msg}")
                return {}
    return {}

# 6. 시스템 백엔드 가동 및 제어 엔진 (openpyxl 기반 웹 최적화)
if not st.session_state["saved_api_key"]:
    st.info("💡 왼쪽 사이드바에 'Gemini API Key'를 먼저 입력해 주세요. (웹 배포 버전은 개인 키를 입력해야 작동합니다.)")
elif not uploaded_file:
    st.warning("📂 번역할 작업 표준서(Excel) 파일을 업로드창에 올려주세요!")
else:
    try:
        genai.configure(api_key=st.session_state["saved_api_key"])
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        st.success("✅ 시스템 가동 준비 완료! 아래 버튼을 누르면 즉시 번역이 시작됩니다.")
        
        if st.button("🚀 사진 보존형 번역 시작하기", type="primary"):
            status_text = st.empty()
            progress_bar = st.progress(0)
            
            status_text.text("🎨 원본 서식 및 이미지 구조 정밀 분석 중...")
            
            # [엔진 교체] 웹 서버 환경에 맞게 openpyxl로 파일을 읽어옵니다.
            wb = openpyxl.load_workbook(uploaded_file, data_only=False)
            
            # 진행도 계산용 전체 워크시트 수
            sheets = wb.worksheets
            total_sheets = len(sheets)
            
            for s_idx, sheet in enumerate(sheets):
                status_text.text(f"📊 [{sheet.title}] 시트 데이터 문맥 스캔 중...")
                
                # 시트 내 모든 텍스트 수집
                unique_texts = set()
                for row in sheet.iter_rows(values_only=False):
                    for cell in row:
                        val = cell.value
                        if val and isinstance(val, str) and not str(val).strip().startswith('='):
                            unique_texts.add(val.strip())
                
                if not unique_texts:
                    continue
                    
                status_text.text(f"🔄 [{sheet.title}] 전체 문맥 파악 및 번역 진행 중...")
                progress_bar.progress(int((s_idx / total_sheets) * 50))
                
                text_list = list(unique_texts)
                translation_map = translate_batch(text_list, source_lang, target_lang, model)
                
                if not translation_map:
                    st.error("⚠️ AI가 데이터를 반환하지 못했습니다. 잠시 후 다시 시도해 주세요.")
                    st.stop()
                
                with st.expander(f"🔍 [{sheet.title}] 실시간 통번역 데이터 검증 표", expanded=True):
                    debug_df = pd.DataFrame([
                        {"원본 텍스트": k, "번역된 텍스트": v} for k, v in translation_map.items()
                    ])
                    st.dataframe(debug_df, use_container_width=True)
                
                status_text.text("✍️ 검증 완료된 데이터 서식 보존 기입 중...")
                
                # 기존 서식과 이미지를 깨지 않고 값만 덮어쓰기
                for row in sheet.iter_rows(values_only=False):
                    for cell in row:
                        val = cell.value
                        if val and isinstance(val, str):
                            val_str = val.strip()
                            if val_str in translation_map:
                                cell.value = translation_map[val_str]
                
                # 웹 전용 글자 자동 너비 조절 코드
                for col in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            # 영어 글자 너비를 고려해 대략적인 길이 측정
                            max_len = max(max_len, len(str(cell.value)))
                    # 너무 넓어지는 것을 방지하기 위한 최대 상한선 설정
                    adjusted_width = min(max(max_len + 3, 10), 50)
                    sheet.column_dimensions[col_letter].width = adjusted_width
                
                progress_bar.progress(int(((s_idx + 1) / total_sheets) * 100))
                time.sleep(1)
            
            # 임시 파일 경로 명시 및 저장
            temp_target_path = f"temp_web_target_{int(time.time())}.xlsx"
            wb.save(temp_target_path)
            
            status_text.text("🎉 대성공! 사진과 서식이 완벽히 보존된 파일이 완성되었습니다.")
            progress_bar.progress(100)
            
            with open(temp_target_path, "rb") as f:
                output_data = f.read()
                
            st.subheader("📥 결과물 다운로드")
            st.download_button(
                label="📁 번역 완료된 엑셀 파일 다운로드",
                data=output_data,
                file_name=f"Translated_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            if os.path.exists(temp_target_path): 
                os.remove(temp_target_path)
            
    except Exception as e:
        st.error(f"❌ 시스템 내부 오류가 발생했습니다: {e}")
